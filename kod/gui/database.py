from PyQt5 import QtCore, QtGui, QtWidgets
from screen import Screen
import openpyxl

class Ui_DatabaseScreen:
    def __init__(self,DatabaseScreen,gui):
        self.gui = gui
        DatabaseScreen.setObjectName("DatabaseScreen")
        DatabaseScreen.resize(781, 878)
        self.centralwidget = QtWidgets.QWidget(DatabaseScreen)
        self.centralwidget.setObjectName("centralwidget")

        self.table_widget = QtWidgets.QTableWidget(self.centralwidget)
        self.table_widget.setGeometry(QtCore.QRect(0,25,1080,850))
        self.load_excel_data()
        
        self.menu = QtWidgets.QPushButton(self.centralwidget)
        self.menu.setGeometry(0,0,40,20)
        self.menu.setText("Menu")
        self.menu.setObjectName("menu")
        self.menu.setStyleSheet("background-color: gray")
        self.menu.clicked.connect(lambda: self.gotoMain())
        
        DatabaseScreen.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(DatabaseScreen)
        self.menubar.setGeometry(QtCore.QRect(0,0,783,26))
        self.menubar.setObjectName("menubar")
        DatabaseScreen.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(DatabaseScreen)
        self.statusbar.setObjectName("statusbar")
        DatabaseScreen.setStatusBar(self.statusbar)
        
        QtCore.QMetaObject.connectSlotsByName(DatabaseScreen)


    def load_excel_data(self):
        try:
            workbook = self.gui.database
            self.gui.stacked_widget.setFixedHeight(876)
            self.gui.stacked_widget.setFixedWidth(1080)
            self.sheet = workbook.active
            self.table_widget.setRowCount(self.sheet.max_row)
            self.table_widget.setColumnCount(self.sheet.max_column)

            for row_index, row in enumerate(self.sheet.iter_rows(values_only=True)):
                for col_index, value in enumerate(row):
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.table_widget.setItem(row_index, col_index, item)



        except Exception as e:
            print(f"Error loading Excel data: {e}")
            
    def gotoMain(self):
        self.gui.stacked_widget.setFixedHeight(876)
        self.gui.stacked_widget.setFixedWidth(783)
        self.gui.show_main()